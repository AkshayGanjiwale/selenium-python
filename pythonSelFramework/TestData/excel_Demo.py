import openpyxl
book = openpyxl.load_workbook("C:\\Users\\ITH\\Documents\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
# reading the value from a cell
cell = sheet.cell(row=1, column=2)
print(cell.value)

# insert value in a cell
sheet.cell(row=2, column =2).value = "Akshay"
print(sheet.cell(row=2, column =2).value)

# identify max. row & column
print(sheet.max_row)

print(sheet.max_column)

# another way to print value of a cell
print(sheet['A5'].value)

# using for loop to print all the values
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)